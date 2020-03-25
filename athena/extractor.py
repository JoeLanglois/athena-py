from openpyxl import load_workbook

class Extractor:
    
    def parse_file(self, path):
        self.wb = load_workbook(path, data_only=True)
        self.sheet = self.wb["Résultats"]
        
        self.names = []
        self.rows = []

        for row in self.sheet.iter_rows():
            row_data = []
            self.names.append(row[0].value)
            
            for cell in row:
                row_data.append(cell.value)
            
            self.rows.append(row_data)

        self.data = self.extract_data()
    
    def get_index_of(self, name):
        return self.names.index(name)  
    
    def get_last_index_of(self, name):
        idxs = []
        for i, row_name in enumerate(self.names):
            if name == row_name:
                idxs.append(i)

        return idxs[len(idxs) - 1]

    def row_to_data(self, rows):
        return  list(((row[0], row[self.current_month_idx]) for row in rows))

    def extract_data(self):
        month_row = self.rows[3]
        rev_start_idx = self.get_index_of("Ventes de services : ") + 1
        rev_end_idx = self.get_index_of("Total des ventes ")

        direct_start_idx = self.get_index_of("Coûts Directs des services:") + 1
        direct_end_idx = self.get_last_index_of("Coûts Directs des services:")

        op_start = self.get_index_of("Frais d'exploitation :") + 1
        op_end = self.get_index_of("Total des frais d'exploitations ")

        self.rev_rows = self.rows[rev_start_idx:rev_end_idx]
        first_rev_row = self.rev_rows[0]

        self.current_month_idx = first_rev_row.index(None) - 1
        current_date = month_row[self.current_month_idx]

        self.rev_data = self.row_to_data(self.rev_rows)  
        self.direct_data = self.row_to_data(self.rows[direct_start_idx:direct_end_idx])    
        self.op_data = self.row_to_data(self.rows[op_start:op_end])

        return {
            "date": current_date,
            "revenues": self.rev_data,
            "direct": self.direct_data,
            "op": self.op_data
        }